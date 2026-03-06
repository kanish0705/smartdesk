"""
Smart Campus Assistant - Simplified Demo
A clean working prototype for project evaluation
"""
from fastapi import FastAPI, Query, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse
from typing import List, Optional
from pydantic import BaseModel
from datetime import date, datetime
import os
import shutil
import uuid
import json

# Try to import openpyxl for Excel parsing
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Create uploads directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
TIMETABLE_DIR = os.path.join(UPLOAD_DIR, "timetables")
ANNOUNCEMENT_DIR = os.path.join(UPLOAD_DIR, "announcements")
RESOURCE_DIR = os.path.join(UPLOAD_DIR, "resources")
STATIC_DIR = os.path.join(os.path.dirname(BASE_DIR), "static")

for directory in [UPLOAD_DIR, TIMETABLE_DIR, ANNOUNCEMENT_DIR, RESOURCE_DIR]:
    os.makedirs(directory, exist_ok=True)

# Initialize FastAPI app
app = FastAPI(
    title="Smart Campus Assistant",
    description="College Student Assistant - Timetable, Announcements & Resources",
    version="1.0.0",
    docs_url="/docs"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== DATA MODELS ====================

class TimetableEntry(BaseModel):
    day: str
    time: str
    subject: str
    room: str
    faculty: str

class Announcement(BaseModel):
    id: int
    title: str
    description: str
    date: str
    priority: str  # high, medium, low
    image_url: Optional[str] = None  # Optional image attachment

class Resource(BaseModel):
    id: int
    title: str
    subject: str
    type: str  # pdf, ppt, link, notes
    url: str
    description: str
    file_name: Optional[str] = None  # Original filename for downloads

# ==================== MOCK DATA ====================

# Timetable data organized by department, semester, section
TIMETABLE_DATA = {
    "BCA": {
        3: {
            "A": [
                {"day": "Monday", "time": "09:00 - 10:00", "subject": "Data Structures", "room": "Lab 1", "faculty": "Dr. Sharma"},
                {"day": "Monday", "time": "10:00 - 11:00", "subject": "Database Systems", "room": "Room 201", "faculty": "Prof. Kumar"},
                {"day": "Monday", "time": "11:15 - 12:15", "subject": "Web Development", "room": "Lab 2", "faculty": "Ms. Singh"},
                {"day": "Monday", "time": "02:00 - 03:00", "subject": "Mathematics", "room": "Room 102", "faculty": "Dr. Patel"},
                {"day": "Tuesday", "time": "09:00 - 10:00", "subject": "Operating Systems", "room": "Room 203", "faculty": "Prof. Gupta"},
                {"day": "Tuesday", "time": "10:00 - 11:00", "subject": "Data Structures Lab", "room": "Lab 1", "faculty": "Dr. Sharma"},
                {"day": "Tuesday", "time": "11:15 - 12:15", "subject": "Data Structures Lab", "room": "Lab 1", "faculty": "Dr. Sharma"},
                {"day": "Tuesday", "time": "02:00 - 03:00", "subject": "English", "room": "Room 105", "faculty": "Ms. Verma"},
                {"day": "Wednesday", "time": "09:00 - 10:00", "subject": "Database Systems", "room": "Room 201", "faculty": "Prof. Kumar"},
                {"day": "Wednesday", "time": "10:00 - 11:00", "subject": "Web Development", "room": "Lab 2", "faculty": "Ms. Singh"},
                {"day": "Wednesday", "time": "11:15 - 12:15", "subject": "Mathematics", "room": "Room 102", "faculty": "Dr. Patel"},
                {"day": "Wednesday", "time": "02:00 - 03:00", "subject": "Operating Systems", "room": "Room 203", "faculty": "Prof. Gupta"},
                {"day": "Thursday", "time": "09:00 - 10:00", "subject": "Data Structures", "room": "Lab 1", "faculty": "Dr. Sharma"},
                {"day": "Thursday", "time": "10:00 - 11:00", "subject": "DBMS Lab", "room": "Lab 3", "faculty": "Prof. Kumar"},
                {"day": "Thursday", "time": "11:15 - 12:15", "subject": "DBMS Lab", "room": "Lab 3", "faculty": "Prof. Kumar"},
                {"day": "Thursday", "time": "02:00 - 03:00", "subject": "English", "room": "Room 105", "faculty": "Ms. Verma"},
                {"day": "Friday", "time": "09:00 - 10:00", "subject": "Web Development", "room": "Lab 2", "faculty": "Ms. Singh"},
                {"day": "Friday", "time": "10:00 - 11:00", "subject": "Operating Systems", "room": "Room 203", "faculty": "Prof. Gupta"},
                {"day": "Friday", "time": "11:15 - 12:15", "subject": "Mathematics", "room": "Room 102", "faculty": "Dr. Patel"},
                {"day": "Friday", "time": "02:00 - 03:00", "subject": "Soft Skills", "room": "Room 101", "faculty": "Mr. Joshi"},
            ],
            "B": [
                {"day": "Monday", "time": "09:00 - 10:00", "subject": "Database Systems", "room": "Room 202", "faculty": "Prof. Kumar"},
                {"day": "Monday", "time": "10:00 - 11:00", "subject": "Data Structures", "room": "Lab 2", "faculty": "Dr. Sharma"},
                {"day": "Monday", "time": "11:15 - 12:15", "subject": "Mathematics", "room": "Room 103", "faculty": "Dr. Patel"},
                {"day": "Monday", "time": "02:00 - 03:00", "subject": "Web Development", "room": "Lab 1", "faculty": "Ms. Singh"},
                {"day": "Tuesday", "time": "09:00 - 10:00", "subject": "English", "room": "Room 106", "faculty": "Ms. Verma"},
                {"day": "Tuesday", "time": "10:00 - 11:00", "subject": "Operating Systems", "room": "Room 204", "faculty": "Prof. Gupta"},
                {"day": "Tuesday", "time": "11:15 - 12:15", "subject": "Data Structures", "room": "Lab 2", "faculty": "Dr. Sharma"},
                {"day": "Tuesday", "time": "02:00 - 03:00", "subject": "Mathematics", "room": "Room 103", "faculty": "Dr. Patel"},
            ]
        }
    },
    "MCA": {
        1: {
            "A": [
                {"day": "Monday", "time": "09:00 - 10:00", "subject": "Programming in Python", "room": "Lab 4", "faculty": "Dr. Reddy"},
                {"day": "Monday", "time": "10:00 - 11:00", "subject": "Computer Networks", "room": "Room 301", "faculty": "Prof. Nair"},
                {"day": "Monday", "time": "11:15 - 12:15", "subject": "Software Engineering", "room": "Room 302", "faculty": "Ms. Iyer"},
                {"day": "Tuesday", "time": "09:00 - 10:00", "subject": "Database Management", "room": "Room 303", "faculty": "Prof. Das"},
                {"day": "Tuesday", "time": "10:00 - 11:00", "subject": "Python Lab", "room": "Lab 4", "faculty": "Dr. Reddy"},
                {"day": "Tuesday", "time": "11:15 - 12:15", "subject": "Python Lab", "room": "Lab 4", "faculty": "Dr. Reddy"},
            ]
        }
    }
}

# Announcements data
ANNOUNCEMENTS_DATA = [
    {
        "id": 1,
        "title": "Mid-Semester Examinations Schedule",
        "description": "Mid-semester examinations will be held from March 15-25, 2026. Students are advised to collect their hall tickets from the examination cell. Detailed timetable will be available on the notice board.",
        "date": "2026-03-05",
        "priority": "high"
    },
    {
        "id": 2,
        "title": "Tech Fest 2026 Registration Open",
        "description": "Annual Tech Fest 'TechnoVerse 2026' registrations are now open. Events include hackathon, coding competition, robotics, and project exhibition. Register before March 20th to avail early bird discount.",
        "date": "2026-03-04",
        "priority": "medium"
    },
    {
        "id": 3,
        "title": "Library Timings Extended",
        "description": "During examination period, library timings have been extended. New timings: 8:00 AM to 10:00 PM (Monday-Saturday). Students can access reading rooms and computer lab facilities.",
        "date": "2026-03-03",
        "priority": "low"
    },
    {
        "id": 4,
        "title": "Guest Lecture on AI & Machine Learning",
        "description": "A guest lecture on 'Future of AI in Industry' will be conducted by Dr. Ramesh from IIT Delhi on March 10th at 2:00 PM in the Main Auditorium. All students are encouraged to attend.",
        "date": "2026-03-02",
        "priority": "medium"
    },
    {
        "id": 5,
        "title": "Campus Placement Drive",
        "description": "Infosys and TCS will be conducting placement drives on March 18th and 20th respectively. Eligible students should register through the placement cell portal before March 12th.",
        "date": "2026-03-01",
        "priority": "high"
    }
]

# Resources data
RESOURCES_DATA = [
    {
        "id": 1,
        "title": "Data Structures Complete Notes",
        "subject": "Data Structures",
        "type": "pdf",
        "url": "https://example.com/ds-notes.pdf",
        "description": "Comprehensive notes covering arrays, linked lists, trees, graphs, and sorting algorithms."
    },
    {
        "id": 2,
        "title": "DBMS Tutorial - W3Schools",
        "subject": "Database Systems",
        "type": "link",
        "url": "https://www.w3schools.com/sql/",
        "description": "Interactive SQL tutorial with examples and practice exercises."
    },
    {
        "id": 3,
        "title": "Web Development Basics",
        "subject": "Web Development",
        "type": "pdf",
        "url": "https://example.com/webdev-basics.pdf",
        "description": "HTML, CSS, and JavaScript fundamentals with practical examples."
    },
    {
        "id": 4,
        "title": "Operating Systems Concepts",
        "subject": "Operating Systems",
        "type": "pdf",
        "url": "https://example.com/os-concepts.pdf",
        "description": "Notes on process management, memory management, and file systems."
    },
    {
        "id": 5,
        "title": "Python Official Documentation",
        "subject": "Programming",
        "type": "link",
        "url": "https://docs.python.org/3/",
        "description": "Official Python 3 documentation and tutorials."
    },
    {
        "id": 6,
        "title": "Computer Networks Notes",
        "subject": "Computer Networks",
        "type": "pdf",
        "url": "https://example.com/cn-notes.pdf",
        "description": "OSI model, TCP/IP, routing protocols, and network security concepts."
    },
    {
        "id": 7,
        "title": "Git & GitHub Tutorial",
        "subject": "Version Control",
        "type": "link",
        "url": "https://guides.github.com/",
        "description": "Learn Git version control and GitHub collaboration."
    },
    {
        "id": 8,
        "title": "Mathematics for Computer Science",
        "subject": "Mathematics",
        "type": "notes",
        "url": "https://example.com/math-cs.pdf",
        "description": "Discrete mathematics, probability, and linear algebra notes."
    }
]

# ==================== API ENDPOINTS ====================

@app.get("/api", tags=["Root"])
async def api_info():
    """API health check and info"""
    return {
        "message": "Welcome to Smart Campus Assistant API",
        "version": "1.0.0",
        "endpoints": ["/timetable", "/announcements", "/resources"],
        "docs": "/docs"
    }

class TimetableCreate(BaseModel):
    department: str
    semester: int
    section: str
    day: str
    time: str
    subject: str
    room: str
    faculty: str

@app.get("/timetable", response_model=List[TimetableEntry], tags=["Timetable"])
async def get_timetable(
    department: str = Query(..., description="Department code (e.g., BCA, MCA)"),
    semester: int = Query(..., description="Semester number (1-8)"),
    section: str = Query(..., description="Section (e.g., A, B)")
):
    """
    Get timetable based on Department, Semester, and Section.
    Returns weekly class schedule.
    """
    dept = department.upper()
    sec = section.upper()
    
    if dept in TIMETABLE_DATA:
        if semester in TIMETABLE_DATA[dept]:
            if sec in TIMETABLE_DATA[dept][semester]:
                return TIMETABLE_DATA[dept][semester][sec]
    
    # Return empty list if no timetable found
    return []

@app.post("/timetable/upload", tags=["Timetable"])
async def upload_timetable_excel(
    department: str = Form(...),
    semester: int = Form(...),
    section: str = Form(...),
    file: UploadFile = File(...)
):
    """
    Upload timetable from Excel file.
    Excel should have columns: Day, Time, Subject, Room, Faculty
    """
    if not EXCEL_SUPPORT:
        raise HTTPException(status_code=500, detail="Excel support not available. Install openpyxl.")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    # Save file temporarily
    file_path = os.path.join(TIMETABLE_DIR, f"{uuid.uuid4()}_{file.filename}")
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    try:
        # Parse Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        entries = []
        headers = [cell.value.lower() if cell.value else "" for cell in sheet[1]]
        
        # Map column indices
        col_map = {}
        for i, h in enumerate(headers):
            if 'day' in h:
                col_map['day'] = i
            elif 'time' in h:
                col_map['time'] = i
            elif 'subject' in h or 'course' in h:
                col_map['subject'] = i
            elif 'room' in h or 'hall' in h:
                col_map['room'] = i
            elif 'faculty' in h or 'teacher' in h or 'professor' in h:
                col_map['faculty'] = i
        
        # Read data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            entry = {
                "day": str(row[col_map.get('day', 0)] or ""),
                "time": str(row[col_map.get('time', 1)] or ""),
                "subject": str(row[col_map.get('subject', 2)] or ""),
                "room": str(row[col_map.get('room', 3)] or ""),
                "faculty": str(row[col_map.get('faculty', 4)] or "")
            }
            if entry["day"] and entry["subject"]:
                entries.append(entry)
        
        # Store in timetable data
        dept = department.upper()
        sec = section.upper()
        
        if dept not in TIMETABLE_DATA:
            TIMETABLE_DATA[dept] = {}
        if semester not in TIMETABLE_DATA[dept]:
            TIMETABLE_DATA[dept][semester] = {}
        
        TIMETABLE_DATA[dept][semester][sec] = entries
        
        return {"message": f"Timetable uploaded successfully. {len(entries)} entries added.", "entries": len(entries)}
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")
    finally:
        # Clean up temp file
        if os.path.exists(file_path):
            os.remove(file_path)

@app.post("/timetable", response_model=TimetableEntry, tags=["Timetable"])
async def add_timetable_entry(timetable: TimetableCreate):
    """
    Add a new timetable entry (Admin functionality).
    """
    dept = timetable.department.upper()
    semester = timetable.semester
    sec = timetable.section.upper()
    
    # Initialize nested dictionaries if they don't exist
    if dept not in TIMETABLE_DATA:
        TIMETABLE_DATA[dept] = {}
    if semester not in TIMETABLE_DATA[dept]:
        TIMETABLE_DATA[dept][semester] = {}
    if sec not in TIMETABLE_DATA[dept][semester]:
        TIMETABLE_DATA[dept][semester][sec] = []
    
    # Create the entry
    entry = {
        "day": timetable.day,
        "time": timetable.time,
        "subject": timetable.subject,
        "room": timetable.room,
        "faculty": timetable.faculty
    }
    
    TIMETABLE_DATA[dept][semester][sec].append(entry)
    
    return entry

@app.get("/announcements", response_model=List[Announcement], tags=["Announcements"])
async def get_announcements():
    """
    Get all announcements.
    Returns announcements sorted by date (newest first).
    """
    return ANNOUNCEMENTS_DATA

@app.post("/announcements", tags=["Announcements"])
async def add_announcement(
    title: str = Form(...),
    description: str = Form(...),
    priority: str = Form("medium"),
    image: Optional[UploadFile] = File(None)
):
    """
    Add a new announcement with optional image (Admin functionality).
    """
    image_url = None
    
    if image and image.filename:
        # Validate image type
        allowed_types = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
        ext = os.path.splitext(image.filename)[1].lower()
        if ext not in allowed_types:
            raise HTTPException(status_code=400, detail=f"Invalid image type. Allowed: {', '.join(allowed_types)}")
        
        # Save image
        unique_filename = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(ANNOUNCEMENT_DIR, unique_filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(image.file, buffer)
        
        image_url = f"/uploads/announcements/{unique_filename}"
    
    new_announcement = {
        "id": int(datetime.now().timestamp() * 1000),
        "title": title,
        "description": description,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "priority": priority,
        "image_url": image_url
    }
    
    ANNOUNCEMENTS_DATA.insert(0, new_announcement)
    return new_announcement

@app.get("/resources", response_model=List[Resource], tags=["Resources"])
async def get_resources(subject: Optional[str] = None):
    """
    Get study resources/materials.
    Optionally filter by subject.
    """
    if subject:
        return [r for r in RESOURCES_DATA if subject.lower() in r["subject"].lower()]
    return RESOURCES_DATA

@app.post("/resources", tags=["Resources"])
async def add_resource(
    title: str = Form(...),
    subject: str = Form(...),
    description: str = Form(...),
    file: UploadFile = File(...)
):
    """
    Upload a resource file (PDF, PPT, etc.) - Admin functionality.
    """
    # Validate file type
    allowed_types = ['.pdf', '.ppt', '.pptx', '.doc', '.docx', '.xls', '.xlsx', '.txt']
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Invalid file type. Allowed: {', '.join(allowed_types)}")
    
    # Determine resource type
    if ext == '.pdf':
        resource_type = 'pdf'
    elif ext in ['.ppt', '.pptx']:
        resource_type = 'ppt'
    elif ext in ['.doc', '.docx']:
        resource_type = 'doc'
    elif ext in ['.xls', '.xlsx']:
        resource_type = 'excel'
    else:
        resource_type = 'file'
    
    # Save file
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = os.path.join(RESOURCE_DIR, unique_filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    new_resource = {
        "id": int(datetime.now().timestamp() * 1000),
        "title": title,
        "subject": subject,
        "type": resource_type,
        "url": f"/uploads/resources/{unique_filename}",
        "description": description,
        "file_name": file.filename
    }
    
    RESOURCES_DATA.insert(0, new_resource)
    return new_resource

# Serve uploaded files
@app.get("/uploads/{folder}/{filename}")
async def serve_upload(folder: str, filename: str):
    """Serve uploaded files"""
    file_path = os.path.join(UPLOAD_DIR, folder, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path)
    raise HTTPException(status_code=404, detail="File not found")

# Health check endpoint for Render
@app.head("/")
@app.get("/health")
async def health_check():
    """Health check for deployment platforms"""
    return {"status": "ok"}

# Serve main UI at root
@app.get("/", response_class=HTMLResponse)
async def serve_index():
    """Serve the student portal"""
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        with open(index_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(content="<h1>Smart Campus Assistant</h1><p>UI files not found. Check deployment.</p>")

# Serve admin panel
@app.get("/admin", response_class=HTMLResponse)
async def serve_admin():
    """Serve the admin panel"""
    admin_path = os.path.join(STATIC_DIR, "admin.html")
    if os.path.exists(admin_path):
        with open(admin_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(content="<h1>Admin Panel</h1><p>Admin UI not found.</p>")

# Mount static files directory for CSS, JS, images
if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
